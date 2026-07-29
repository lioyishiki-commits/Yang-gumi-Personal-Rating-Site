from __future__ import annotations

import io
import unittest
from datetime import datetime

from openpyxl import load_workbook

import analysis_export
import scoring


class AnalysisExportTest(unittest.TestCase):
    def test_xlsx_is_styled_acgn_wide_and_contains_every_score(self):
        config = scoring.default_score_config()
        works = [
            {
                "id": 1,
                "title": "自动动画",
                "original_title": "AUTO ANIME",
                "type": "动画",
                "subtype": "TV",
                "status": "已看",
                "year": 2026,
                "release_date": "2026-04-01",
                "score_mode": "auto",
                "score_total": 8.14,
                "bangumi_score": 8.30,
                "bangumi_rank": 83,
                "bangumi_total_votes": 8690,
                "score_imbalance_penalty": 0.5,
                "short_review": "自动评分测试",
                "private_note": "绝不能导出",
                "resource_path": r"E:\private\video.mkv",
                "score_breakdown": {
                    "body": {field: 8.1 for field in scoring.score_weights("body", config)},
                    "feeling": {field: 8.2 for field in scoring.score_weights("feeling", config)},
                    "era": {field: 8.3 for field in scoring.score_weights("era", config)},
                },
            },
            {
                "id": 2,
                "title": "手动动画",
                "type": "动画",
                "score_mode": "manual",
                "score_total": 8.2,
                "bangumi_score": 7.95,
                "score_breakdown": {},
            },
            {
                "id": 3,
                "title": "游戏也必须出现",
                "type": "游戏",
                "score_mode": "manual",
                "score_total": 9.0,
                "bangumi_score": 8.0,
            },
        ]
        payload = analysis_export.build_readonly_analysis_xlsx(
            works,
            config,
            generated_at=datetime(2026, 7, 26, 12, 0, 0),
        )

        self.assertTrue(payload.startswith(b"PK"))
        workbook = load_workbook(io.BytesIO(payload), data_only=False)
        self.assertEqual(workbook.sheetnames, ["导出说明", "评分明细", "评分规则"])
        detail = workbook["评分明细"]
        headers = {cell.value: cell.column for cell in detail[3]}
        expected = {
            "类型", "我的总评分", "Bangumi 评分", "评分差", "剧情", "角色塑造",
            "作画 / 摄影", "演出", "音乐 / 配音", "节奏", "个人偏爱",
            "重看 / 重玩价值", "情绪后劲", "独特性", "氛围感", "影响力",
            "开创性", "偏科惩罚",
        }
        self.assertTrue(expected.issubset(headers))
        self.assertEqual(detail.max_row, 6)
        title_column = headers["作品名"]
        rows = {
            detail.cell(row, title_column).value: row
            for row in range(4, detail.max_row + 1)
        }
        self.assertEqual(set(rows), {"自动动画", "手动动画", "游戏也必须出现"})
        auto_row = rows["自动动画"]
        manual_row = rows["手动动画"]
        self.assertEqual(detail.cell(auto_row, headers["评分差"]).value, -0.16)
        self.assertEqual(detail.cell(auto_row, headers["我的总评分"]).number_format, "0.00")
        self.assertEqual(detail.cell(manual_row, headers["我的总评分"]).number_format, "0.0")
        self.assertEqual(detail.cell(auto_row, headers["Bangumi 评分"]).number_format, "0.00")
        self.assertEqual(detail["A3"].value, "序号")
        self.assertEqual(detail["B3"].value, "作品名")
        self.assertEqual(detail.freeze_panes, "C4")
        self.assertEqual(detail.auto_filter.ref.split(":")[0], "A3")
        self.assertGreaterEqual(detail.column_dimensions["B"].width, 38)
        self.assertTrue(detail.cell(auto_row, title_column).alignment.wrap_text)
        self.assertEqual(detail.cell(auto_row, headers["序号"]).alignment.horizontal, "center")
        self.assertFalse(any(
            merged.min_col <= 2 < merged.max_col and merged.min_row <= 2
            for merged in detail.merged_cells.ranges
        ))
        self.assertEqual(detail["A2"].value, "共 3 部作品")
        self.assertIn("Bangumi 评分仅供对照", detail["C2"].value)
        all_text = " ".join(
            str(cell.value or "")
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
        self.assertNotIn("绝不能导出", all_text)
        self.assertNotIn(r"E:\private", all_text)
        self.assertIn("游戏也必须出现", all_text)


if __name__ == "__main__":
    unittest.main()

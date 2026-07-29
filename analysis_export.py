"""Privacy-safe, styled XLSX export for owner and read-only analysis pages."""
from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

import database as db
import scoring


PINK = "E95D8B"
INK = "13151A"
PANEL = "20232A"
MUTED = "858B96"
WHITE = "F4F4F6"
CYAN = "63CBE7"
GOLD = "E5BB63"
LINE = "343840"
SOFT = "F3F4F7"
PALE_PINK = "FCE8EF"


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _component_value(work: dict[str, Any], group: str, field: str) -> Any:
    breakdown = work.get("score_breakdown")
    if isinstance(breakdown, dict):
        values = breakdown.get(group)
        if isinstance(values, dict) and field in values:
            return values.get(field)
    if field in work:
        return work.get(field)
    try:
        custom = json.loads(work.get("custom_scores_json") or "{}")
    except (TypeError, ValueError, json.JSONDecodeError):
        custom = {}
    return custom.get(field) if isinstance(custom, dict) else None


def _score_difference(work: dict[str, Any]) -> float | None:
    mine = _number(work.get("score_total"))
    public = _number(work.get("bangumi_score"))
    return None if mine is None or public is None else round(mine - public, 2)


def _style_title(ws, title: str, subtitle: str, end_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=INK)
    cell.font = Font(name="Microsoft YaHei UI", size=18, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    cell = ws.cell(2, 1, subtitle)
    cell.fill = PatternFill("solid", fgColor=PANEL)
    cell.font = Font(name="Microsoft YaHei UI", size=10, color=MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 24


def _style_detail_title(
    ws,
    *,
    item_count: int,
    generated_at: datetime,
    end_column: int,
) -> None:
    """Style the detail heading without merging across the B/C freeze line."""
    title_fill = PatternFill("solid", fgColor=INK)
    subtitle_fill = PatternFill("solid", fgColor=PANEL)
    for column in range(1, end_column + 1):
        ws.cell(1, column).fill = title_fill
        ws.cell(2, column).fill = subtitle_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws["A1"] = "逐部评分明细"
    ws["A1"].font = Font(name="Microsoft YaHei UI", size=16, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=end_column)
    ws["C1"] = "YANG·GUMI / ACGN SCORE ARCHIVE"
    ws["C1"].font = Font(name="Microsoft YaHei UI", size=10, bold=True, color=PINK)
    ws["C1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws["A2"] = f"共 {item_count} 部作品"
    ws["A2"].font = Font(name="Microsoft YaHei UI", size=10, bold=True, color=WHITE)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=end_column)
    ws["C2"] = f"Bangumi 评分仅供对照 · 生成于 {generated_at:%Y-%m-%d %H:%M:%S}"
    ws["C2"].font = Font(name="Microsoft YaHei UI", size=10, color=MUTED)
    ws["C2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 24


def _style_table(ws, header_row: int, first_data_row: int, last_row: int, last_column: int) -> None:
    thin = Side(style="thin", color=LINE)
    for cell in ws[header_row]:
        if cell.column > last_column:
            break
        cell.fill = PatternFill("solid", fgColor=PINK)
        cell.font = Font(name="Microsoft YaHei UI", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.row_dimensions[header_row].height = 30
    for row in range(first_data_row, last_row + 1):
        fill = PatternFill("solid", fgColor="FFFFFF" if row % 2 else SOFT)
        for column in range(1, last_column + 1):
            cell = ws.cell(row, column)
            cell.fill = fill
            cell.font = Font(name="Microsoft YaHei UI", size=9, color=INK)
            cell.alignment = Alignment(
                horizontal="left" if column <= 3 else "center",
                vertical="center",
                wrap_text=column in {2, 3, last_column},
            )
            cell.border = Border(bottom=Side(style="hair", color="D9DCE2"))
        ws.row_dimensions[row].height = 28


def _set_widths(ws, widths: Iterable[float]) -> None:
    for index, width in enumerate(widths, 1):
        letter = get_column_letter(index)
        ws.column_dimensions[letter] = ColumnDimension(ws, min=index, max=index, width=width)


def _average(rows: list[dict[str, Any]], field: str) -> float | None:
    values = [_number(row.get(field)) for row in rows]
    present = [value for value in values if value is not None]
    return round(sum(present) / len(present), 2) if present else None


def build_readonly_analysis_xlsx(
    works: Iterable[dict[str, Any]] | None = None,
    config: dict[str, Any] | None = None,
    generated_at: datetime | None = None,
) -> bytes:
    """Build an XLSX containing privacy-safe ACGN totals and every score item."""
    rows = [dict(row) for row in (works if works is not None else db.public_rows())]
    records = list(rows)
    records.sort(key=lambda row: (
        _number(row.get("score_total")) is not None,
        _number(row.get("score_total")) or -1,
        str(row.get("title") or ""),
    ), reverse=True)
    active_config = config or scoring.load_score_config()
    generated_at = generated_at or datetime.now()

    component_columns: list[tuple[str, str, str]] = []
    for group in ("body", "feeling", "era"):
        labels = scoring.score_labels(group, active_config)
        for field in scoring.score_weights(group, active_config):
            component_columns.append((group, field, labels.get(field, field)))

    workbook = Workbook()
    workbook.remove(workbook.active)
    overview = workbook.create_sheet("导出说明")
    detail = workbook.create_sheet("评分明细")
    rules = workbook.create_sheet("评分规则")

    _style_title(
        overview,
        "YANG·GUMI ACGN 评分明细",
        "主站与只读分享导出 · 不含私人备注、本地路径、备份与维护数据",
        6,
    )
    overview_rows = [
        ("导出时间", generated_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("作品条目", len(records)),
        ("已评分", sum(_number(row.get("score_total")) is not None for row in records)),
        ("自动综合", sum(row.get("score_mode") == "auto" for row in records)),
        ("手动总评", sum(row.get("score_mode") == "manual" for row in records)),
        ("我的平均", _average(records, "score_total")),
        ("Bangumi 平均", _average(records, "bangumi_score")),
    ]
    overview.append([])
    overview.append(["项目", "结果", "说明", "", "", ""])
    for label, value in overview_rows:
        explanation = {
            "自动综合": "自动综合评分保留两位小数",
            "手动总评": "手动总评分显示一位小数",
            "我的平均": "所有已评分作品的平均值",
            "Bangumi 平均": "仅作公开评分参考",
        }.get(label, "")
        overview.append([label, value, explanation, "", "", ""])
    overview["D4"] = "文件内容"
    overview["D5"] = "评分明细：逐部列出类型、总评、Bangumi 评分、差值与所有评分小项目。"
    overview["D6"] = "评分规则：列出当前评分组上限、项目权重与单项最高贡献。"
    overview["D8"] = "精度规则"
    overview["D9"] = "手动总评分 1 位小数；自动综合、Bangumi、差值、分项及统计均为 2 位小数。"
    overview.merge_cells("D5:F6")
    overview.merge_cells("D9:F11")
    for cell in ("D4", "D8"):
        overview[cell].font = Font(name="Microsoft YaHei UI", bold=True, color=PINK)
    for cell in ("D5", "D9"):
        overview[cell].alignment = Alignment(wrap_text=True, vertical="top")
        overview[cell].font = Font(name="Microsoft YaHei UI", color=INK)
    _style_table(overview, 4, 5, 11, 3)
    _set_widths(overview, [18, 20, 34, 18, 18, 18])
    overview.freeze_panes = "A5"
    for row in range(9, 12):
        overview.cell(row, 2).number_format = "0.00"

    base_headers = [
        "序号", "作品名", "原名", "类型", "子类型", "状态", "年份", "首播 / 发售日期", "评分方式",
        "我的总评分", "Bangumi 评分", "评分差", "Bangumi 排名", "评分人数",
    ]
    headers = base_headers + [label for _, _, label in component_columns] + ["偏科惩罚", "一句话短评"]
    _style_detail_title(
        detail,
        item_count=len(records),
        generated_at=generated_at,
        end_column=len(headers),
    )
    detail.append(headers)
    for index, work in enumerate(records, 1):
        score_mode = "自动综合" if work.get("score_mode") == "auto" else "手动总评" if work.get("score_mode") == "manual" else "旧记录"
        values = [
            index,
            work.get("title") or "",
            work.get("original_title") or "",
            work.get("type") or "",
            work.get("subtype") or "",
            "抛弃" if work.get("status") == "弃置" else work.get("status") or "",
            work.get("year"),
            work.get("release_date") or "",
            score_mode,
            _number(work.get("score_total")),
            _number(work.get("bangumi_score")),
            _score_difference(work),
            work.get("bangumi_rank"),
            work.get("bangumi_total_votes"),
        ]
        values.extend(_number(_component_value(work, group, field)) for group, field, _ in component_columns)
        values.extend([
            _number(work.get("score_imbalance_penalty")),
            work.get("short_review") or "",
        ])
        detail.append(values)
        row_number = detail.max_row
        detail.cell(row_number, 10).number_format = "0.0" if work.get("score_mode") == "manual" else "0.00"
        for column in range(11, len(headers)):
            if column not in {13, 14}:
                detail.cell(row_number, column).number_format = "0.00"
        detail.cell(row_number, 12).number_format = "+0.00;-0.00;0.00"
        detail.cell(row_number, 14).number_format = "#,##0"
    last_detail_row = max(detail.max_row, 3)
    _style_table(detail, 3, 4, last_detail_row, len(headers))
    for row in range(4, last_detail_row + 1):
        detail.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        detail.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        detail.cell(row, 3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        detail.cell(row, len(headers)).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
        detail.row_dimensions[row].height = 38
    detail.freeze_panes = "C4"
    detail.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{last_detail_row}"
    _set_widths(
        detail,
        [7, 40, 36, 10, 11, 10, 9, 15, 12, 13, 14, 11, 13, 14]
        + [14] * len(component_columns)
        + [12, 34],
    )

    rule_headers = ["评分组", "组上限", "项目", "项目权重", "单项最高贡献", "字段"]
    _style_title(
        rules,
        "当前自动综合评分规则",
        "权重取自当前评分设置；手动总评不参与自动计算",
        len(rule_headers),
    )
    rules.append(rule_headers)
    for group in ("body", "feeling", "era"):
        cap = scoring.score_cap(group, active_config)
        labels = scoring.score_labels(group, active_config)
        group_label = scoring.SCORE_GROUPS[group]
        for field, weight in scoring.score_weights(group, active_config).items():
            rules.append([
                group_label,
                cap,
                labels.get(field, field),
                weight,
                round(cap * float(weight), 4),
                field,
            ])
    _style_table(rules, 3, 4, rules.max_row, len(rule_headers))
    rules.freeze_panes = "A4"
    rules.auto_filter.ref = f"A3:F{rules.max_row}"
    _set_widths(rules, [16, 13, 24, 14, 18, 29])
    for row in range(4, rules.max_row + 1):
        rules.cell(row, 2).number_format = "0.00"
        rules.cell(row, 4).number_format = "0.00%"
        rules.cell(row, 5).number_format = "0.0000"

    workbook.properties.title = "Yang-gumi ACGN 评分明细"
    workbook.properties.subject = "主站与只读分享评分明细及当前评分规则"
    workbook.properties.creator = "Yang-gumi"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
